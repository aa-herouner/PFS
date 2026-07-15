"""Export a report dict to Excel (openpyxl) or PDF (xhtml2pdf)."""
import io

from django.http import HttpResponse
from django.template.loader import render_to_string


def _filename(report, ext):
    slug = report['title'].lower().replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '')
    return f'{slug}.{ext}'


def to_excel(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = report['title'][:31]  # sheet name limit

    ws.append([report['title']])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(report['headers'])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for row in report['rows']:
        ws.append([_cell(v) for v in row])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{_filename(report, "xlsx")}"'
    return resp


def _cell(value):
    # openpyxl can't write Decimal-in-str cleanly for some types; keep simple.
    if value is None:
        return ''
    return value


def to_pdf(report):
    from xhtml2pdf import pisa

    html = render_to_string('reports/report_pdf.html', {'report': report})
    buf = io.BytesIO()
    result = pisa.CreatePDF(src=html, dest=buf)
    if result.err:
        return HttpResponse('Error generating PDF', status=500)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{_filename(report, "pdf")}"'
    return resp
